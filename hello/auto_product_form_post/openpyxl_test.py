from openpyxl import load_workbook
import time


def read_xlsx(read_only):
    t0 = time.time()
    wb = load_workbook(r'lawcdyc200m.xlsx', read_only=read_only)
    t1 = time.time()
    print(wb.sheetnames)
    # print(sh.cell(row=1, column=1).value)
    # print(sh.cell(row=100, column=3).value)
    print('耗时%0.3f秒钟' % (t1 - t0))



read_xlsx(False)
read_xlsx(True)


# from openpyxl import Workbook
#
# wb = Workbook()
# sh = wb.active
# sh.append(['id', '语文', '数学', '英语', '物理'])
# for i in range(1000000):  # 写入100万行数据
#     sh.append([i + 1, 90, 100, 95, 99])
#
# wb.save(r'bigxlsx.xlsx')
